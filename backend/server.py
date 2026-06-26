from dotenv import load_dotenv
from pathlib import Path
from urllib.parse import urlparse

BASE_DIR = Path(__file__).resolve().parent
REPO_ROOT = BASE_DIR.parent
load_dotenv(REPO_ROOT / ".env")
load_dotenv(BASE_DIR / ".env", override=False)

from fastapi import FastAPI, APIRouter, HTTPException, Request, Response, Depends, Query, WebSocket, WebSocketDisconnect, UploadFile, File
from fastapi.responses import StreamingResponse
from fastapi.staticfiles import StaticFiles
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
from pydantic import BaseModel, Field
from typing import List, Optional
from datetime import datetime, timezone, timedelta
from bson import ObjectId
import os
import logging
import uuid
import bcrypt
import jwt
import io
import json
import asyncio
import requests as http_requests
try:
    from emergentintegrations.llm.chat import LlmChat, UserMessage
except ImportError:
    LlmChat = None
    UserMessage = None
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors as rl_colors
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from openpyxl import Workbook
from openpyxl.styles import Font as XlFont, Alignment, PatternFill
from pd_routes import pd_router, init_pd, run_stability_scheduler, check_stability_alerts_for_tenant
from crm_routes import crm_router, init_crm, run_alert_scheduler
from estoque_routes import estoque_router, init_estoque
from recebimento_routes import recebimento_router, init_recebimento
from retrabalho_routes import retrabalho_router, init_retrabalho
from expedicao_routes import expedicao_router, init_expedicao
from faturamento_routes import faturamento_router, init_faturamento
from pcp_routes import pcp_router, init_pcp
from orders_routes import orders_router, ops_router, init_orders
from kickoff_routes import kickoff_router, init_kickoff
from compras_routes import compras_router, init_compras, create_compras_indexes
from contratos_routes import contratos_router, init_contratos
from cq_routes import cq_router, init_cq, create_cq_indexes
from categorias_routes import categorias_router, init_categorias, create_categorias_indexes
from fragrancias_routes import fragrancias_router, init_fragrancias, create_fragrancias_indexes
from materiais_routes import materiais_router, init_materiais, create_materiais_indexes
from produtos_routes import produtos_router, init_produtos, create_produtos_indexes
from propostas_routes import propostas_router, init_propostas
from requirements_routes import requirements_router, init_requirements, create_requirements_indexes
from workflow_engine import init_workflow, run_workflow_notification_scheduler
from workflow_routes import workflow_router, init_workflow_routes
from rbac import (
    require_roles,
    has_role,
    COMERCIAL_FULL,
    PD_FULL,
    PD_READ,
    ADMIN_ONLY,
    DOC_REVIEWERS,
)

# Roles allowed to use the legacy commercial pipeline (cards, stages, fields, amostras, products, messages, AI helpers, proposal)
COMMERCIAL_PIPELINE_ROLES = COMERCIAL_FULL | {"admin"}
# Roles allowed to read the commercial board (vendedor, sales_ops, sucesso_cliente, admin). Sales ops/manager can also read; PD roles are blocked.
COMMERCIAL_PIPELINE_READ_ROLES = COMMERCIAL_PIPELINE_ROLES

# ============ SETUP ============
missing_env_defaults = []

mongo_url = os.environ.get("MONGO_URL", "mongodb://127.0.0.1:27017")
if "MONGO_URL" not in os.environ:
    missing_env_defaults.append("MONGO_URL")

db_name = os.environ.get("DB_NAME", "kuryos_crm")
if "DB_NAME" not in os.environ:
    missing_env_defaults.append("DB_NAME")

client = AsyncIOMotorClient(mongo_url)
db = client[db_name]

app = FastAPI(title="CRM Kuryos API")
router = APIRouter(prefix="/api")


@app.middleware("http")
async def force_utf8_json_responses(request: Request, call_next):
    response = await call_next(request)
    content_type = response.headers.get("content-type", "")
    if "application/json" in content_type.lower() and "charset=" not in content_type.lower():
        response.headers["content-type"] = "application/json; charset=utf-8"
    return response

JWT_SECRET = os.environ.get("JWT_SECRET", "dev-only-change-me")
if "JWT_SECRET" not in os.environ:
    missing_env_defaults.append("JWT_SECRET")
JWT_ALGORITHM = "HS256"

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

if missing_env_defaults:
    logger.warning(
        "Using local development defaults for: %s",
        ", ".join(missing_env_defaults),
    )

# ============ HELPERS ============

def ensure_ai_available():
    if LlmChat is None or UserMessage is None:
        raise HTTPException(
            status_code=503,
            detail="Recursos de IA indisponiveis no ambiente local: dependencia 'emergentintegrations' nao instalada."
        )

def new_id():
    return str(uuid.uuid4())

def now_iso():
    return datetime.now(timezone.utc).isoformat()

def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")

def verify_password(plain: str, hashed: str) -> bool:
    return bcrypt.checkpw(plain.encode("utf-8"), hashed.encode("utf-8"))

def create_access_token(user_id: str, email: str, tenant_id: str, role: str) -> str:
    payload = {
        "sub": user_id, "email": email, "tenant_id": tenant_id, "role": role,
        "exp": datetime.now(timezone.utc) + timedelta(hours=12), "type": "access"
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)

def create_refresh_token(user_id: str) -> str:
    payload = {"sub": user_id, "exp": datetime.now(timezone.utc) + timedelta(days=7), "type": "refresh"}
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)

async def get_current_user(request: Request) -> dict:
    token = request.cookies.get("access_token")
    if not token:
        auth = request.headers.get("Authorization", "")
        if auth.startswith("Bearer "):
            token = auth[7:]
    if not token:
        raise HTTPException(status_code=401, detail="Not authenticated")
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        if payload.get("type") != "access":
            raise HTTPException(status_code=401, detail="Invalid token type")
        user = await db.users.find_one({"id": payload["sub"]})
        if not user:
            raise HTTPException(status_code=401, detail="User not found")
        return {k: v for k, v in user.items() if k not in ("_id", "password_hash")}
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")

_IS_PRODUCTION = bool(os.environ.get("RENDER") or os.environ.get("ENVIRONMENT") == "production")
_COOKIE_SECURE = _IS_PRODUCTION
_COOKIE_SAMESITE = "none" if _IS_PRODUCTION else "lax"

def set_auth_cookies(response: Response, access: str, refresh: str):
    response.set_cookie(key="access_token", value=access, httponly=True, secure=_COOKIE_SECURE, samesite=_COOKIE_SAMESITE, max_age=43200, path="/")
    response.set_cookie(key="refresh_token", value=refresh, httponly=True, secure=_COOKIE_SECURE, samesite=_COOKIE_SAMESITE, max_age=604800, path="/")


async def get_commercial_user(request: Request) -> dict:
    """Auth + ensure user belongs to commercial pipeline roles."""
    user = await get_current_user(request)
    require_roles(user, COMMERCIAL_PIPELINE_ROLES)
    return user

# ============ WEBSOCKET MANAGER ============

class ConnectionManager:
    def __init__(self):
        self.active_connections: dict = {}

    async def connect(self, websocket: WebSocket, tenant_id: str):
        await websocket.accept()
        if tenant_id not in self.active_connections:
            self.active_connections[tenant_id] = []
        self.active_connections[tenant_id].append(websocket)

    def disconnect(self, websocket: WebSocket, tenant_id: str):
        if tenant_id in self.active_connections:
            self.active_connections[tenant_id] = [c for c in self.active_connections[tenant_id] if c != websocket]

    async def broadcast(self, tenant_id: str, event: str, data: dict):
        if tenant_id not in self.active_connections:
            return
        message = json.dumps({"event": event, "data": data})
        disconnected = []
        for conn in self.active_connections.get(tenant_id, []):
            try:
                await conn.send_text(message)
            except Exception:
                disconnected.append(conn)
        for conn in disconnected:
            if conn in self.active_connections.get(tenant_id, []):
                self.active_connections[tenant_id].remove(conn)

ws_manager = ConnectionManager()

# ============ OBJECT STORAGE ============

STORAGE_URL = "https://integrations.emergentagent.com/objstore/api/v1/storage"
APP_NAME = "kuryos-crm"
storage_key = None

def init_storage():
    global storage_key
    if storage_key:
        return storage_key
    key = os.environ.get("EMERGENT_LLM_KEY")
    if not key:
        logger.warning("No EMERGENT_LLM_KEY for storage")
        return None
    try:
        resp = http_requests.post(f"{STORAGE_URL}/init", json={"emergent_key": key}, timeout=30)
        resp.raise_for_status()
        storage_key = resp.json()["storage_key"]
        logger.info("Object storage initialized")
        return storage_key
    except Exception as e:
        logger.error(f"Storage init failed: {e}")
        return None

def put_object(path: str, data: bytes, content_type: str) -> dict:
    key = init_storage()
    if not key:
        raise HTTPException(status_code=500, detail="Storage not initialized")
    resp = http_requests.put(
        f"{STORAGE_URL}/objects/{path}",
        headers={"X-Storage-Key": key, "Content-Type": content_type},
        data=data, timeout=120
    )
    resp.raise_for_status()
    return resp.json()

def get_object(path: str):
    key = init_storage()
    if not key:
        raise HTTPException(status_code=500, detail="Storage not initialized")
    resp = http_requests.get(
        f"{STORAGE_URL}/objects/{path}",
        headers={"X-Storage-Key": key}, timeout=60
    )
    resp.raise_for_status()
    return resp.content, resp.headers.get("Content-Type", "application/octet-stream")

# ============ PYDANTIC MODELS ============

class RegisterInput(BaseModel):
    email: str
    password: str
    name: str
    org_name: str

class LoginInput(BaseModel):
    email: str
    password: str

class CardCreate(BaseModel):
    stage_id: str
    pipeline_id: str
    nome_cliente: str
    telefone: str = ""
    email: str = ""
    status: str = "frio"
    produto: str = ""
    nome_projeto: str = ""
    objetivo_projeto: str = ""
    aplicacoes_desenvolver: str = ""
    ativos_claims: str = ""
    referencias: str = ""
    referencias_fotos_url: str = ""
    orcamento_projeto: str = ""
    textura_esperada: str = ""
    aplicacao: str = ""
    sensorial: str = ""
    ph: str = ""
    outras_observacoes: str = ""

class CardUpdate(BaseModel):
    nome_cliente: Optional[str] = None
    telefone: Optional[str] = None
    email: Optional[str] = None
    status: Optional[str] = None
    produto: Optional[str] = None
    nome_projeto: Optional[str] = None
    objetivo_projeto: Optional[str] = None
    aplicacoes_desenvolver: Optional[str] = None
    ativos_claims: Optional[str] = None
    referencias: Optional[str] = None
    referencias_fotos_url: Optional[str] = None
    orcamento_projeto: Optional[str] = None
    textura_esperada: Optional[str] = None
    aplicacao: Optional[str] = None
    sensorial: Optional[str] = None
    ph: Optional[str] = None
    outras_observacoes: Optional[str] = None

class CardMove(BaseModel):
    stage_id: str
    justification: Optional[str] = None

class FieldValueSave(BaseModel):
    field_id: str
    value_json: str

class ProductCreate(BaseModel):
    nome_produto: str
    sku: str = ""
    quantidade: int = 1
    valor_unitario: float = 0.0

class TaskCreate(BaseModel):
    card_id: Optional[str] = None
    titulo: str
    data_vencimento: Optional[str] = None
    status: str = "pendente"
    assignee_id: Optional[str] = None

class TaskUpdate(BaseModel):
    titulo: Optional[str] = None
    data_vencimento: Optional[str] = None
    status: Optional[str] = None

class MessageCreate(BaseModel):
    content: str
    msg_type: str = "text"

class InviteInput(BaseModel):
    email: str
    name: str
    role: str = "vendedor"

class RoleUpdate(BaseModel):
    role: str

class UserPatch(BaseModel):
    name: Optional[str] = None
    new_password: Optional[str] = None

# ============ AUTH ROUTES ============

@router.post("/auth/register")
async def register(input_data: RegisterInput, response: Response):
    email = input_data.email.lower().strip()
    existing = await db.users.find_one({"email": email})
    if existing:
        raise HTTPException(status_code=400, detail="Email already registered")

    tenant_id = new_id()
    await db.tenants.insert_one({"id": tenant_id, "name": input_data.org_name, "created_at": now_iso()})

    user_id = new_id()
    await db.users.insert_one({
        "id": user_id, "email": email, "password_hash": hash_password(input_data.password),
        "name": input_data.name, "role": "admin", "tenant_id": tenant_id, "created_at": now_iso()
    })

    await seed_default_pipeline(tenant_id)

    access = create_access_token(user_id, email, tenant_id, "admin")
    refresh = create_refresh_token(user_id)
    set_auth_cookies(response, access, refresh)
    return {"id": user_id, "email": email, "name": input_data.name, "role": "admin", "tenant_id": tenant_id}

@router.post("/auth/login")
async def login(input_data: LoginInput, response: Response):
    email = input_data.email.lower().strip()
    user = await db.users.find_one({"email": email})
    if not user:
        raise HTTPException(status_code=401, detail="Invalid credentials")
    if not verify_password(input_data.password, user["password_hash"]):
        raise HTTPException(status_code=401, detail="Invalid credentials")

    access = create_access_token(user["id"], email, user["tenant_id"], user["role"])
    refresh = create_refresh_token(user["id"])
    set_auth_cookies(response, access, refresh)
    return {"id": user["id"], "email": email, "name": user["name"], "role": user["role"], "tenant_id": user["tenant_id"]}

@router.post("/auth/logout")
async def logout(response: Response):
    response.delete_cookie("access_token", path="/")
    response.delete_cookie("refresh_token", path="/")
    return {"message": "Logged out"}

@router.get("/auth/me")
async def me(request: Request):
    user = await get_current_user(request)
    return user

@router.post("/auth/refresh")
async def refresh_token(request: Request, response: Response):
    token = request.cookies.get("refresh_token")
    if not token:
        raise HTTPException(status_code=401, detail="No refresh token")
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        if payload.get("type") != "refresh":
            raise HTTPException(status_code=401, detail="Invalid token type")
        user = await db.users.find_one({"id": payload["sub"]})
        if not user:
            raise HTTPException(status_code=401, detail="User not found")
        access = create_access_token(user["id"], user["email"], user["tenant_id"], user["role"])
        response.set_cookie(key="access_token", value=access, httponly=True, secure=False, samesite="lax", max_age=43200, path="/")
        return {"message": "Token refreshed"}
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Refresh token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid refresh token")

# ============ PIPELINE / BOARD ROUTES ============

@router.get("/pipelines")
async def list_pipelines(request: Request):
    user = await get_commercial_user(request)
    pipelines = await db.pipelines.find({"tenant_id": user["tenant_id"]}, {"_id": 0}).to_list(100)
    return pipelines

@router.get("/pipelines/{pipeline_id}/board")
async def get_board(pipeline_id: str, request: Request):
    user = await get_commercial_user(request)
    pipeline = await db.pipelines.find_one({"id": pipeline_id, "tenant_id": user["tenant_id"]}, {"_id": 0})
    if not pipeline:
        raise HTTPException(status_code=404, detail="Pipeline not found")

    stages = await db.stages.find({"pipeline_id": pipeline_id}, {"_id": 0}).sort("order", 1).to_list(100)
    stage_ids = [s["id"] for s in stages]

    fields = await db.fields.find({"stage_id": {"$in": stage_ids}}, {"_id": 0}).to_list(500)
    cards = await db.cards.find({"pipeline_id": pipeline_id, "tenant_id": user["tenant_id"]}, {"_id": 0}).to_list(5000)

    users_list = await db.users.find({"tenant_id": user["tenant_id"]}, {"_id": 0, "password_hash": 0}).to_list(100)

    fields_by_stage = {}
    for f in fields:
        fields_by_stage.setdefault(f["stage_id"], []).append(f)

    cards_by_stage = {}
    for c in cards:
        cards_by_stage.setdefault(c["stage_id"], []).append(c)

    board_stages = []
    for s in stages:
        board_stages.append({
            **s,
            "fields": fields_by_stage.get(s["id"], []),
            "cards": cards_by_stage.get(s["id"], [])
        })

    return {"pipeline": pipeline, "stages": board_stages, "users": users_list}

# ============ STAGE MANAGEMENT ============

class StageCreate(BaseModel):
    pipeline_id: str
    name: str
    order: Optional[int] = None

class StageUpdate(BaseModel):
    name: Optional[str] = None
    order: Optional[int] = None

class StageReorder(BaseModel):
    stage_ids: List[str]

class FieldCreate(BaseModel):
    stage_id: str
    label: str
    type: str = "text"
    required: bool = False
    options: List[str] = []

class FieldUpdate(BaseModel):
    label: Optional[str] = None
    type: Optional[str] = None
    required: Optional[bool] = None
    options: Optional[List[str]] = None

@router.post("/stages")
async def create_stage(data: StageCreate, request: Request):
    user = await get_commercial_user(request)
    pipeline = await db.pipelines.find_one({"id": data.pipeline_id, "tenant_id": user["tenant_id"]})
    if not pipeline:
        raise HTTPException(status_code=404, detail="Pipeline não encontrado")

    if data.order is None:
        max_order = await db.stages.find({"pipeline_id": data.pipeline_id}).sort("order", -1).to_list(1)
        data.order = (max_order[0]["order"] + 1) if max_order else 0

    stage_id = new_id()
    stage = {
        "id": stage_id, "pipeline_id": data.pipeline_id,
        "name": data.name, "order": data.order, "created_at": now_iso()
    }
    await db.stages.insert_one(stage)
    stage.pop("_id", None)
    return stage

@router.put("/stages/{stage_id}")
async def update_stage(stage_id: str, data: StageUpdate, request: Request):
    user = await get_commercial_user(request)
    stage = await db.stages.find_one({"id": stage_id}, {"_id": 0})
    if not stage:
        raise HTTPException(status_code=404, detail="Estágio não encontrado")
    pipeline = await db.pipelines.find_one({"id": stage["pipeline_id"], "tenant_id": user["tenant_id"]})
    if not pipeline:
        raise HTTPException(status_code=403, detail="Sem permissão")

    updates = {k: v for k, v in data.model_dump(exclude_unset=True).items() if v is not None}
    if not updates:
        raise HTTPException(status_code=400, detail="Nada para atualizar")

    await db.stages.update_one({"id": stage_id}, {"$set": updates})
    updated = await db.stages.find_one({"id": stage_id}, {"_id": 0})
    return updated

@router.delete("/stages/{stage_id}")
async def delete_stage(stage_id: str, request: Request):
    user = await get_commercial_user(request)
    stage = await db.stages.find_one({"id": stage_id}, {"_id": 0})
    if not stage:
        raise HTTPException(status_code=404, detail="Estágio não encontrado")
    pipeline = await db.pipelines.find_one({"id": stage["pipeline_id"], "tenant_id": user["tenant_id"]})
    if not pipeline:
        raise HTTPException(status_code=403, detail="Sem permissão")

    cards_count = await db.cards.count_documents({"stage_id": stage_id})
    if cards_count > 0:
        raise HTTPException(status_code=400, detail=f"Não é possível excluir: {cards_count} card(s) neste estágio. Mova-os primeiro.")

    await db.fields.delete_many({"stage_id": stage_id})
    await db.stages.delete_one({"id": stage_id})
    return {"message": "Estágio removido"}

@router.put("/stages/reorder")
async def reorder_stages(data: StageReorder, request: Request):
    user = await get_commercial_user(request)
    for i, sid in enumerate(data.stage_ids):
        await db.stages.update_one({"id": sid}, {"$set": {"order": i}})
    return {"message": "Estágios reordenados"}

@router.post("/fields")
async def create_field(data: FieldCreate, request: Request):
    user = await get_commercial_user(request)
    stage = await db.stages.find_one({"id": data.stage_id}, {"_id": 0})
    if not stage:
        raise HTTPException(status_code=404, detail="Estágio não encontrado")

    field_id = new_id()
    field = {
        "id": field_id, "stage_id": data.stage_id,
        "label": data.label, "type": data.type,
        "required": data.required, "options": data.options,
        "created_at": now_iso()
    }
    await db.fields.insert_one(field)
    field.pop("_id", None)
    return field

@router.put("/fields/{field_id}")
async def update_field(field_id: str, data: FieldUpdate, request: Request):
    user = await get_commercial_user(request)
    updates = {k: v for k, v in data.model_dump(exclude_unset=True).items() if v is not None}
    if not updates:
        raise HTTPException(status_code=400, detail="Nada para atualizar")

    result = await db.fields.update_one({"id": field_id}, {"$set": updates})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Campo não encontrado")
    updated = await db.fields.find_one({"id": field_id}, {"_id": 0})
    return updated

@router.delete("/fields/{field_id}")
async def delete_field(field_id: str, request: Request):
    user = await get_commercial_user(request)
    field = await db.fields.find_one({"id": field_id})
    if not field:
        raise HTTPException(status_code=404, detail="Campo não encontrado")

    await db.field_values.delete_many({"field_id": field_id})
    await db.fields.delete_one({"id": field_id})
    return {"message": "Campo removido"}

# ============ CARD ROUTES ============

@router.post("/cards")
async def create_card(data: CardCreate, request: Request):
    user = await get_commercial_user(request)
    card_id = new_id()
    card = {
        "id": card_id, "tenant_id": user["tenant_id"], "pipeline_id": data.pipeline_id,
        "stage_id": data.stage_id, "nome_cliente": data.nome_cliente, "telefone": data.telefone,
        "email": data.email, "vendedor_id": user["id"], "status": data.status,
        "produto": data.produto, "nome_projeto": data.nome_projeto,
        "objetivo_projeto": data.objetivo_projeto, "aplicacoes_desenvolver": data.aplicacoes_desenvolver,
        "ativos_claims": data.ativos_claims, "referencias": data.referencias,
        "referencias_fotos_url": data.referencias_fotos_url, "orcamento_projeto": data.orcamento_projeto,
        "textura_esperada": data.textura_esperada, "aplicacao": data.aplicacao,
        "sensorial": data.sensorial, "ph": data.ph, "outras_observacoes": data.outras_observacoes,
        "created_at": now_iso()
    }
    await db.cards.insert_one(card)
    await db.card_history.insert_one({
        "id": new_id(), "card_id": card_id, "action": "Card criado",
        "details": f"Lead {data.nome_cliente} criado na fase", "user_id": user["id"],
        "user_name": user["name"], "created_at": now_iso()
    })
    card.pop("_id", None)
    await ws_manager.broadcast(user["tenant_id"], "card_created", {"card": card})
    return card

@router.get("/cards/{card_id}")
async def get_card(card_id: str, request: Request):
    user = await get_commercial_user(request)
    card = await db.cards.find_one({"id": card_id, "tenant_id": user["tenant_id"]}, {"_id": 0})
    if not card:
        raise HTTPException(status_code=404, detail="Card not found")
    return card

@router.put("/cards/{card_id}")
async def update_card(card_id: str, data: CardUpdate, request: Request):
    user = await get_commercial_user(request)
    update_fields = {k: v for k, v in data.model_dump().items() if v is not None}
    if not update_fields:
        raise HTTPException(status_code=400, detail="No fields to update")

    result = await db.cards.update_one(
        {"id": card_id, "tenant_id": user["tenant_id"]},
        {"$set": update_fields}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Card not found")

    changes = ", ".join(f"{k}: {v}" for k, v in update_fields.items())
    await db.card_history.insert_one({
        "id": new_id(), "card_id": card_id, "action": "Card atualizado",
        "details": changes, "user_id": user["id"], "user_name": user["name"], "created_at": now_iso()
    })

    card = await db.cards.find_one({"id": card_id}, {"_id": 0})

    # ---- AUTOMATION: Hot lead notification ----
    if update_fields.get("status") == "quente":
        notif = {
            "id": new_id(), "tenant_id": user["tenant_id"],
            "user_id": card.get("vendedor_id", user["id"]),
            "type": "hot_lead", "title": "Lead Quente!",
            "message": f"O lead {card['nome_cliente']} foi marcado como QUENTE!",
            "card_id": card_id, "read": False, "created_at": now_iso()
        }
        await db.notifications.insert_one(notif)
        notif.pop("_id", None)
        await ws_manager.broadcast(user["tenant_id"], "notification", notif)
        await db.email_logs.insert_one({
            "id": new_id(), "tenant_id": user["tenant_id"],
            "to": user["email"], "subject": f"Lead QUENTE - {card['nome_cliente']}",
            "body": f"O lead {card['nome_cliente']} foi marcado como quente. Priorize o contato!",
            "status": "mock_sent", "created_at": now_iso()
        })

    return card

@router.put("/cards/{card_id}/move")
async def move_card(card_id: str, data: CardMove, request: Request):
    user = await get_commercial_user(request)
    card = await db.cards.find_one({"id": card_id, "tenant_id": user["tenant_id"]}, {"_id": 0})
    if not card:
        raise HTTPException(status_code=404, detail="Card not found")

    old_stage = await db.stages.find_one({"id": card["stage_id"]}, {"_id": 0})
    new_stage = await db.stages.find_one({"id": data.stage_id}, {"_id": 0})

    is_backward = (
        old_stage and new_stage and
        new_stage.get("order", 0) < old_stage.get("order", 0)
    )
    if is_backward and not (data.justification or "").strip():
        raise HTTPException(
            status_code=422,
            detail="Justificativa obrigatória para mover um lead para fase anterior."
        )

    await db.cards.update_one({"id": card_id}, {"$set": {"stage_id": data.stage_id}})

    old_name = old_stage["name"] if old_stage else "?"
    new_name = new_stage["name"] if new_stage else "?"
    action_label = "Retorno de fase (requer revisão do líder)" if is_backward else "Movido de fase"
    details = f"{old_name} → {new_name}"
    if is_backward and data.justification:
        details += f" | Justificativa: {data.justification.strip()}"
    await db.card_history.insert_one({
        "id": new_id(), "card_id": card_id, "action": action_label,
        "details": details, "user_id": user["id"],
        "user_name": user["name"], "created_at": now_iso(),
        "is_backward": is_backward,
    })

    if is_backward:
        # Create a lider review task
        lider = await db.users.find_one(
            {"tenant_id": user["tenant_id"], "role": {"$in": ["admin", "gestor", "sales_ops"]}},
            {"_id": 0, "id": 1, "name": 1}
        )
        task_assignee_id = lider["id"] if lider else user["id"]
        task_assignee_name = lider["name"] if lider else user["name"]
        await db.tasks.insert_one({
            "id": new_id(), "tenant_id": user["tenant_id"],
            "title": f"Revisar retorno de fase: {card.get('nome_cliente', card_id)}",
            "description": f"O lead '{card.get('nome_cliente', '')}' foi movido de volta para '{new_name}' (antes: '{old_name}').\nJustificativa: {(data.justification or '').strip()}\nSolicitado por: {user['name']}",
            "assignee_id": task_assignee_id, "assignee_name": task_assignee_name,
            "created_by_id": user["id"], "created_by_name": user["name"],
            "status": "pendente", "priority": "alta",
            "entity_type": "card", "entity_id": card_id,
            "created_at": now_iso(), "due_date": None,
        })


    updated = await db.cards.find_one({"id": card_id}, {"_id": 0})
    await ws_manager.broadcast(user["tenant_id"], "card_moved", {"card": updated, "from_stage": old_name, "to_stage": new_name})

    # ---- AUTOMATION: Negociando Proposta ----
    if new_stage and "negociando" in new_stage["name"].lower():
        notif = {
            "id": new_id(), "tenant_id": user["tenant_id"],
            "user_id": updated.get("vendedor_id", user["id"]),
            "type": "automation", "title": "Proposta automatica sugerida",
            "message": f"O lead {updated['nome_cliente']} entrou em {new_name}. Gere uma proposta PDF!",
            "card_id": card_id, "read": False, "created_at": now_iso()
        }
        await db.notifications.insert_one(notif)
        notif.pop("_id", None)
        await ws_manager.broadcast(user["tenant_id"], "notification", notif)
        # Mock email log
        await db.email_logs.insert_one({
            "id": new_id(), "tenant_id": user["tenant_id"],
            "to": user["email"], "subject": f"Lead {updated['nome_cliente']} - Proposta sugerida",
            "body": f"O lead {updated['nome_cliente']} foi movido para {new_name}. Acesse o CRM para gerar a proposta.",
            "status": "mock_sent", "created_at": now_iso()
        })

    # ---- AUTOMATION: Amostras -> Criar solicitação P&D POR AMOSTRA ----
    if new_stage and "amostra" in new_stage["name"].lower():
        existing_pd = await db.pd_requests.find_one({
            "client_card_id": card_id, "tenant_id": user["tenant_id"]
        })
        if not existing_pd:
            # Determine priority
            fv_cursor = db.field_values.find({"card_id": card_id}, {"_id": 0})
            field_vals = await fv_cursor.to_list(100)
            priority = "Normal"
            for fv in field_vals:
                val = fv.get("value_json", "")
                if val in ("Alta", "Alto"):
                    priority = "Alta"
                    break
                elif val in ("Baixa", "Baixo"):
                    priority = "Baixa"
                    break
                elif val == "Urgente":
                    priority = "Urgente"
                    break
            
            # Build briefing from card fields
            briefing_parts = []
            if updated.get("objetivo_projeto"):
                briefing_parts.append(f"Objetivo: {updated['objetivo_projeto']}")
            if updated.get("aplicacoes_desenvolver"):
                briefing_parts.append(f"Aplicações: {updated['aplicacoes_desenvolver']}")
            if updated.get("ativos_claims"):
                briefing_parts.append(f"Ativos/Claims: {updated['ativos_claims']}")
            if updated.get("textura_esperada"):
                briefing_parts.append(f"Textura: {updated['textura_esperada']}")
            if updated.get("aplicacao"):
                briefing_parts.append(f"Aplicação: {updated['aplicacao']}")
            if updated.get("sensorial"):
                briefing_parts.append(f"Sensorial: {updated['sensorial']}")
            if updated.get("ph"):
                briefing_parts.append(f"pH: {updated['ph']}")
            if updated.get("orcamento_projeto"):
                briefing_parts.append(f"Orçamento: {updated['orcamento_projeto']}")
            if updated.get("outras_observacoes"):
                briefing_parts.append(f"Observações: {updated['outras_observacoes']}")
            base_description = "\n".join(briefing_parts) if briefing_parts else ""
            
            # Get amostras for this card
            card_amostras = await db.card_amostras.find(
                {"card_id": card_id, "tenant_id": user["tenant_id"]}, {"_id": 0}
            ).to_list(100)
            
            created_count = 0
            
            if card_amostras:
                # Create ONE PD request per amostra
                for amostra in card_amostras:
                    pd_id = new_id()
                    amostra_desc = []
                    if amostra.get("tipo_produto"):
                        amostra_desc.append(f"Tipo: {amostra['tipo_produto']}")
                    if amostra.get("descricao"):
                        amostra_desc.append(f"Descrição: {amostra['descricao']}")
                    if amostra.get("referencia"):
                        amostra_desc.append(f"Referência: {amostra['referencia']}")
                    if amostra.get("volume"):
                        amostra_desc.append(f"Volume: {amostra['volume']}")
                    if amostra.get("observacoes"):
                        amostra_desc.append(f"Obs: {amostra['observacoes']}")
                    
                    full_description = base_description
                    if amostra_desc:
                        full_description += ("\n\n--- Dados da Amostra ---\n" + "\n".join(amostra_desc)) if full_description else "\n".join(amostra_desc)
                    
                    pd_request = {
                        "id": pd_id,
                        "tenant_id": user["tenant_id"],
                        "client_card_id": card_id,
                        "amostra_id": amostra["id"],
                        "client_name": updated.get("nome_cliente", ""),
                        "project_name": amostra["nome_amostra"],
                        "request_type": amostra.get("tipo_produto") or "Produto Novo",
                        "category": "",
                        "description": full_description,
                        "references": amostra.get("referencia") or updated.get("referencias", ""),
                        "restrictions": "",
                        "volume": amostra.get("volume") or "",
                        "packaging": "",
                        "priority": priority,
                        "deadline": None,
                        "status": "OPEN",
                        "sku": "",
                        "created_by": user["id"],
                        "created_by_name": user["name"],
                        "created_at": now_iso(),
                        "updated_at": now_iso(),
                    }
                    await db.pd_requests.insert_one(pd_request)
                    await db.pd_request_status_history.insert_one({
                        "id": new_id(), "pd_request_id": pd_id,
                        "from_status": None, "to_status": "OPEN",
                        "changed_by": user["id"], "changed_by_name": user["name"],
                        "comment": f"Amostra '{amostra['nome_amostra']}' de '{updated['nome_cliente']}' — Criada automaticamente",
                        "created_at": now_iso(),
                    })
                    created_count += 1
            else:
                # Fallback: no amostras defined, create 1 PD request
                pd_id = new_id()
                project_name = updated.get("nome_projeto") or updated.get("produto") or f"Projeto {updated['nome_cliente']}"
                pd_request = {
                    "id": pd_id,
                    "tenant_id": user["tenant_id"],
                    "client_card_id": card_id,
                    "amostra_id": None,
                    "client_name": updated.get("nome_cliente", ""),
                    "project_name": project_name,
                    "request_type": "Produto Novo",
                    "category": "",
                    "description": base_description,
                    "references": updated.get("referencias", ""),
                    "restrictions": "",
                    "volume": "",
                    "packaging": "",
                    "priority": priority,
                    "deadline": None,
                    "status": "OPEN",
                    "sku": "",
                    "created_by": user["id"],
                    "created_by_name": user["name"],
                    "created_at": now_iso(),
                    "updated_at": now_iso(),
                }
                await db.pd_requests.insert_one(pd_request)
                await db.pd_request_status_history.insert_one({
                    "id": new_id(), "pd_request_id": pd_id,
                    "from_status": None, "to_status": "OPEN",
                    "changed_by": user["id"], "changed_by_name": user["name"],
                    "comment": f"Criado automaticamente para '{updated['nome_cliente']}' — Briefing anexado",
                    "created_at": now_iso(),
                })
                created_count = 1
            
            notif_pd = {
                "id": new_id(), "tenant_id": user["tenant_id"],
                "user_id": user["id"],
                "type": "automation", "title": f"{created_count} amostra(s) P&D criada(s)",
                "message": f"{created_count} solicitação(ões) P&D criada(s) para '{updated['nome_cliente']}'. Acesse P&D para gerenciar.",
                "card_id": card_id, "read": False, "created_at": now_iso()
            }
            await db.notifications.insert_one(notif_pd)
            notif_pd.pop("_id", None)
            await ws_manager.broadcast(user["tenant_id"], "notification", notif_pd)
            logger.info(f"Auto-created {created_count} PD request(s) for card {card_id}")

    return updated

@router.delete("/cards/{card_id}")
async def delete_card(card_id: str, request: Request):
    user = await get_commercial_user(request)
    result = await db.cards.delete_one({"id": card_id, "tenant_id": user["tenant_id"]})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Card not found")
    await db.field_values.delete_many({"card_id": card_id})
    await db.card_products.delete_many({"card_id": card_id})
    await db.card_amostras.delete_many({"card_id": card_id})
    await db.card_history.delete_many({"card_id": card_id})
    await db.messages.delete_many({"card_id": card_id})
    return {"message": "Card deleted"}

# ============ CARD DETAILS (aggregate) ============

@router.get("/cards/{card_id}/details")
async def get_card_details(card_id: str, request: Request):
    user = await get_commercial_user(request)
    card = await db.cards.find_one({"id": card_id, "tenant_id": user["tenant_id"]}, {"_id": 0})
    if not card:
        raise HTTPException(status_code=404, detail="Card not found")

    stage = await db.stages.find_one({"id": card["stage_id"]}, {"_id": 0})
    fields = await db.fields.find({"stage_id": card["stage_id"]}, {"_id": 0}).to_list(50)
    field_values = await db.field_values.find({"card_id": card_id}, {"_id": 0}).to_list(100)
    products = await db.card_products.find({"card_id": card_id}, {"_id": 0}).to_list(100)
    history = await db.card_history.find({"card_id": card_id}, {"_id": 0}).sort("created_at", -1).to_list(200)
    messages = await db.messages.find({"card_id": card_id}, {"_id": 0}).sort("created_at", 1).to_list(500)
    amostras = await db.card_amostras.find({"card_id": card_id}, {"_id": 0}).sort("created_at", 1).to_list(100)

    return {
        "card": card, "stage": stage, "fields": fields,
        "field_values": field_values, "products": products,
        "history": history, "messages": messages, "amostras": amostras
    }

# ============ CARD AMOSTRAS (Sample Requests) ============

class AmostraCreate(BaseModel):
    nome_amostra: str
    tipo_produto: str = ""
    descricao: str = ""
    referencia: str = ""
    volume: str = ""
    observacoes: str = ""

class AmostraUpdate(BaseModel):
    nome_amostra: Optional[str] = None
    tipo_produto: Optional[str] = None
    descricao: Optional[str] = None
    referencia: Optional[str] = None
    volume: Optional[str] = None
    observacoes: Optional[str] = None

@router.post("/cards/{card_id}/amostras")
async def add_amostra(card_id: str, data: AmostraCreate, request: Request):
    user = await get_commercial_user(request)
    card = await db.cards.find_one({"id": card_id, "tenant_id": user["tenant_id"]})
    if not card:
        raise HTTPException(status_code=404, detail="Card not found")
    amostra_id = new_id()
    amostra = {
        "id": amostra_id,
        "card_id": card_id,
        "tenant_id": user["tenant_id"],
        "nome_amostra": data.nome_amostra,
        "tipo_produto": data.tipo_produto,
        "descricao": data.descricao,
        "referencia": data.referencia,
        "volume": data.volume,
        "observacoes": data.observacoes,
        "created_at": now_iso(),
    }
    await db.card_amostras.insert_one(amostra)
    amostra.pop("_id", None)
    return amostra

@router.get("/cards/{card_id}/amostras")
async def list_amostras(card_id: str, request: Request):
    user = await get_commercial_user(request)
    amostras = await db.card_amostras.find(
        {"card_id": card_id, "tenant_id": user["tenant_id"]}, {"_id": 0}
    ).sort("created_at", 1).to_list(100)
    return amostras

@router.put("/cards/{card_id}/amostras/{amostra_id}")
async def update_amostra(card_id: str, amostra_id: str, data: AmostraUpdate, request: Request):
    user = await get_commercial_user(request)
    update_fields = {k: v for k, v in data.model_dump().items() if v is not None}
    if not update_fields:
        raise HTTPException(status_code=400, detail="No fields to update")
    result = await db.card_amostras.update_one(
        {"id": amostra_id, "card_id": card_id, "tenant_id": user["tenant_id"]},
        {"$set": update_fields}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Amostra not found")
    amostra = await db.card_amostras.find_one({"id": amostra_id}, {"_id": 0})
    return amostra

@router.delete("/cards/{card_id}/amostras/{amostra_id}")
async def delete_amostra(card_id: str, amostra_id: str, request: Request):
    user = await get_commercial_user(request)
    result = await db.card_amostras.delete_one(
        {"id": amostra_id, "card_id": card_id, "tenant_id": user["tenant_id"]}
    )
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Amostra not found")
    return {"message": "Amostra removida"}

# ============ FIELD VALUES ============

@router.post("/cards/{card_id}/field-values")
async def save_field_values(card_id: str, values: List[FieldValueSave], request: Request):
    user = await get_commercial_user(request)
    card = await db.cards.find_one({"id": card_id, "tenant_id": user["tenant_id"]})
    if not card:
        raise HTTPException(status_code=404, detail="Card not found")

    for v in values:
        existing = await db.field_values.find_one({"card_id": card_id, "field_id": v.field_id})
        if existing:
            await db.field_values.update_one(
                {"card_id": card_id, "field_id": v.field_id},
                {"$set": {"value_json": v.value_json}}
            )
        else:
            await db.field_values.insert_one({
                "id": new_id(), "card_id": card_id, "field_id": v.field_id, "value_json": v.value_json
            })

    await db.card_history.insert_one({
        "id": new_id(), "card_id": card_id, "action": "Campos atualizados",
        "details": f"{len(values)} campo(s) salvo(s)", "user_id": user["id"],
        "user_name": user["name"], "created_at": now_iso()
    })

    return {"message": "Field values saved"}

# ============ CARD PRODUCTS ============

@router.get("/cards/{card_id}/products")
async def list_products(card_id: str, request: Request):
    await get_commercial_user(request)
    products = await db.card_products.find({"card_id": card_id}, {"_id": 0}).to_list(100)
    return products

@router.post("/cards/{card_id}/products")
async def add_product(card_id: str, data: ProductCreate, request: Request):
    user = await get_commercial_user(request)
    product = {
        "id": new_id(), "card_id": card_id, "nome_produto": data.nome_produto,
        "sku": data.sku, "quantidade": data.quantidade, "valor_unitario": data.valor_unitario,
        "valor_total": data.quantidade * data.valor_unitario, "created_at": now_iso()
    }
    await db.card_products.insert_one(product)
    product.pop("_id", None)

    await db.card_history.insert_one({
        "id": new_id(), "card_id": card_id, "action": "Produto adicionado",
        "details": f"{data.nome_produto} (x{data.quantidade})", "user_id": user["id"],
        "user_name": user["name"], "created_at": now_iso()
    })

    return product

@router.delete("/card-products/{product_id}")
async def delete_product(product_id: str, request: Request):
    await get_commercial_user(request)
    result = await db.card_products.delete_one({"id": product_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Product not found")
    return {"message": "Product deleted"}

# ============ TASKS ============

@router.get("/tasks")
async def list_tasks(request: Request):
    user = await get_current_user(request)
    tasks = await db.tasks.find({"tenant_id": user["tenant_id"]}, {"_id": 0}).sort("created_at", -1).to_list(500)
    return tasks

@router.post("/tasks")
async def create_task(data: TaskCreate, request: Request):
    user = await get_current_user(request)
    task = {
        "id": new_id(), "tenant_id": user["tenant_id"], "card_id": data.card_id,
        "titulo": data.titulo, "data_vencimento": data.data_vencimento,
        "status": data.status, "assignee_id": data.assignee_id or user["id"],
        "created_at": now_iso()
    }
    await db.tasks.insert_one(task)
    task.pop("_id", None)
    return task

@router.put("/tasks/{task_id}")
async def update_task(task_id: str, data: TaskUpdate, request: Request):
    user = await get_current_user(request)
    update_fields = {k: v for k, v in data.model_dump().items() if v is not None}
    if not update_fields:
        raise HTTPException(status_code=400, detail="No fields to update")
    result = await db.tasks.update_one(
        {"id": task_id, "tenant_id": user["tenant_id"]},
        {"$set": update_fields}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Task not found")
    task = await db.tasks.find_one({"id": task_id}, {"_id": 0})
    return task

@router.delete("/tasks/{task_id}")
async def delete_task(task_id: str, request: Request):
    user = await get_current_user(request)
    result = await db.tasks.delete_one({"id": task_id, "tenant_id": user["tenant_id"]})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Task not found")
    return {"message": "Task deleted"}

# ============ MESSAGES (Mock WhatsApp) ============

@router.get("/cards/{card_id}/messages")
async def list_messages(card_id: str, request: Request):
    await get_commercial_user(request)
    messages = await db.messages.find({"card_id": card_id}, {"_id": 0}).sort("created_at", 1).to_list(500)
    return messages

@router.post("/cards/{card_id}/messages")
async def send_message(card_id: str, data: MessageCreate, request: Request):
    user = await get_commercial_user(request)
    message = {
        "id": new_id(), "card_id": card_id, "tenant_id": user["tenant_id"],
        "content": data.content, "msg_type": data.msg_type, "sender": "agent",
        "sender_name": user["name"], "created_at": now_iso()
    }
    await db.messages.insert_one(message)
    message.pop("_id", None)
    return message

    await db.messages.insert_one(message)
    message.pop("_id", None)
    return message

# ============ USER MANAGEMENT + RBAC ============

@router.post("/users/invite")
async def invite_user(data: InviteInput, request: Request):
    user = await get_current_user(request)
    if user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Somente admins podem convidar usuarios")

    email = data.email.lower().strip()
    existing = await db.users.find_one({"email": email})
    if existing:
        raise HTTPException(status_code=400, detail="Email ja registrado")

    valid_roles = ("admin", "vendedor", "sales_ops", "formulador", "qa", "lider_pd", "engenharia_produto", "sucesso_cliente", "gestor")
    if data.role not in valid_roles:
        raise HTTPException(status_code=400, detail=f"Role invalida. Use: {', '.join(valid_roles)}")

    temp_password = f"Kuryos{uuid.uuid4().hex[:6]}!"
    new_user = {
        "id": new_id(), "email": email, "password_hash": hash_password(temp_password),
        "name": data.name, "role": data.role, "tenant_id": user["tenant_id"],
        "invited_by": user["id"], "created_at": now_iso()
    }
    await db.users.insert_one(new_user)

    # Mock email with temp password
    await db.email_logs.insert_one({
        "id": new_id(), "tenant_id": user["tenant_id"],
        "to": email, "subject": "Convite para CRM Kuryos",
        "body": f"Voce foi convidado para o CRM Kuryos.\nEmail: {email}\nSenha temporaria: {temp_password}",
        "status": "mock_sent", "created_at": now_iso()
    })

    # Notification
    await db.notifications.insert_one({
        "id": new_id(), "tenant_id": user["tenant_id"], "user_id": user["id"],
        "type": "user_invited", "title": "Usuario convidado",
        "message": f"{data.name} ({email}) foi convidado como {data.role}",
        "card_id": None, "read": False, "created_at": now_iso()
    })

    logger.info(f"User invited: {email} with temp password: {temp_password}")
    return {"id": new_user["id"], "email": email, "name": data.name, "role": data.role, "temp_password": temp_password}

@router.put("/users/{user_id}/role")
async def update_user_role(user_id: str, data: RoleUpdate, request: Request):
    user = await get_current_user(request)
    if user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Somente admins podem alterar roles")
    valid_roles = ("admin", "vendedor", "sales_ops", "formulador", "qa", "lider_pd", "engenharia_produto", "sucesso_cliente", "gestor")
    if data.role not in valid_roles:
        raise HTTPException(status_code=400, detail=f"Role invalida. Use: {', '.join(valid_roles)}")
    if user_id == user["id"]:
        raise HTTPException(status_code=400, detail="Nao pode alterar sua propria role")

    result = await db.users.update_one(
        {"id": user_id, "tenant_id": user["tenant_id"]},
        {"$set": {"role": data.role}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Usuario nao encontrado")
    return {"message": "Role atualizada", "user_id": user_id, "role": data.role}

@router.patch("/users/{user_id}")
async def update_user(user_id: str, data: UserPatch, request: Request):
    user = await get_current_user(request)
    if user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Somente admins podem editar usuarios")
    patch: dict = {}
    if data.name:
        patch["name"] = data.name.strip()
    if data.new_password:
        if len(data.new_password) < 6:
            raise HTTPException(status_code=400, detail="Senha deve ter no minimo 6 caracteres")
        patch["password_hash"] = hash_password(data.new_password)
    if not patch:
        raise HTTPException(status_code=400, detail="Nenhum campo para atualizar")
    result = await db.users.update_one(
        {"id": user_id, "tenant_id": user["tenant_id"]},
        {"$set": patch}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Usuario nao encontrado")
    return {"message": "Usuario atualizado", "user_id": user_id}

@router.delete("/users/{user_id}")
async def remove_user(user_id: str, request: Request):
    user = await get_current_user(request)
    if user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Somente admins podem remover usuarios")
    if user_id == user["id"]:
        raise HTTPException(status_code=400, detail="Nao pode remover a si mesmo")

    result = await db.users.delete_one({"id": user_id, "tenant_id": user["tenant_id"]})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Usuario nao encontrado")
    return {"message": "Usuario removido"}

# ============ FILE UPLOAD / DOWNLOAD ============

MIME_TYPES = {
    "jpg": "image/jpeg", "jpeg": "image/jpeg", "png": "image/png",
    "gif": "image/gif", "webp": "image/webp", "pdf": "application/pdf",
    "csv": "text/csv", "txt": "text/plain", "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "docx": "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
}

@router.post("/upload")
async def upload_file(request: Request, file: UploadFile = File(...)):
    user = await get_current_user(request)
    data = await file.read()
    if len(data) > 10 * 1024 * 1024:  # 10MB limit
        raise HTTPException(status_code=400, detail="Arquivo muito grande (max 10MB)")

    ext = file.filename.split(".")[-1].lower() if "." in file.filename else "bin"
    ct = file.content_type or MIME_TYPES.get(ext, "application/octet-stream")
    path = f"{APP_NAME}/uploads/{user['tenant_id']}/{uuid.uuid4()}.{ext}"

    result = put_object(path, data, ct)

    file_doc = {
        "id": new_id(), "tenant_id": user["tenant_id"],
        "storage_path": result["path"], "original_filename": file.filename,
        "content_type": ct, "size": result.get("size", len(data)),
        "uploaded_by": user["id"], "is_deleted": False, "created_at": now_iso()
    }
    await db.files.insert_one(file_doc)
    file_doc.pop("_id", None)
    return file_doc

@router.get("/files/{file_id}")
async def download_file(file_id: str, request: Request):
    user = await get_current_user(request)
    record = await db.files.find_one({"id": file_id, "tenant_id": user["tenant_id"], "is_deleted": False}, {"_id": 0})
    if not record:
        raise HTTPException(status_code=404, detail="Arquivo nao encontrado")

    data, ct = get_object(record["storage_path"])
    return Response(content=data, media_type=record.get("content_type", ct),
        headers={"Content-Disposition": f'inline; filename="{record["original_filename"]}"'})

# ============ NOTIFICATIONS ============

@router.get("/notifications")
async def list_notifications(request: Request):
    user = await get_current_user(request)
    notifs = await db.notifications.find(
        {"tenant_id": user["tenant_id"], "user_id": user["id"]},
        {"_id": 0}
    ).sort("created_at", -1).to_list(50)
    return notifs

@router.get("/notifications/count")
async def unread_notification_count(request: Request):
    user = await get_current_user(request)
    count = await db.notifications.count_documents(
        {"tenant_id": user["tenant_id"], "user_id": user["id"], "read": False}
    )
    return {"count": count}

@router.put("/notifications/{notif_id}/read")
async def mark_notification_read(notif_id: str, request: Request):
    user = await get_current_user(request)
    await db.notifications.update_one(
        {"id": notif_id, "tenant_id": user["tenant_id"]},
        {"$set": {"read": True}}
    )
    return {"message": "Notificacao lida"}

@router.put("/notifications/read-all")
async def mark_all_read(request: Request):
    user = await get_current_user(request)
    await db.notifications.update_many(
        {"tenant_id": user["tenant_id"], "user_id": user["id"], "read": False},
        {"$set": {"read": True}}
    )
    return {"message": "Todas lidas"}

# ============ WHATSAPP TEMPLATES ============

WHATSAPP_TEMPLATES = [
    {"id": "welcome", "name": "Boas-vindas", "content": "Ola {nome}! Sou {vendedor} da {empresa}. Gostaria de apresentar nossas solucoes em cosmeticos e perfumaria. Quando seria um bom momento para conversarmos?"},
    {"id": "followup", "name": "Follow-up", "content": "Ola {nome}! Estou entrando em contato para dar continuidade a nossa conversa. Tem alguma duvida sobre os produtos que apresentamos?"},
    {"id": "proposal", "name": "Proposta Enviada", "content": "Ola {nome}! Enviei a proposta comercial conforme conversamos. Por favor, analise e me avise se tiver alguma duvida. Estou a disposicao!"},
    {"id": "sample", "name": "Envio de Amostras", "content": "Ola {nome}! As amostras solicitadas foram enviadas. Em breve voce recebera e podera avaliar a qualidade dos nossos produtos."},
    {"id": "closing", "name": "Fechamento", "content": "Ola {nome}! Parabens pela escolha! Estamos muito felizes em ter voce como parceiro. Vamos alinhar os proximos passos?"},
]

@router.get("/whatsapp/templates")
async def get_templates(request: Request):
    await get_current_user(request)
    return WHATSAPP_TEMPLATES

# ============ EMAIL LOGS (Mock) ============

@router.get("/email-logs")
async def list_email_logs(request: Request):
    user = await get_current_user(request)
    logs = await db.email_logs.find(
        {"tenant_id": user["tenant_id"]}, {"_id": 0}
    ).sort("created_at", -1).to_list(50)
    return logs

# ============ AI ENDPOINTS (Claude Sonnet 4.5) ============

@router.post("/ai/lead-summary/{card_id}")
async def ai_lead_summary(card_id: str, request: Request):
    ensure_ai_available()
    user = await get_commercial_user(request)
    card = await db.cards.find_one({"id": card_id, "tenant_id": user["tenant_id"]}, {"_id": 0})
    if not card:
        raise HTTPException(status_code=404, detail="Card not found")

    stage = await db.stages.find_one({"id": card["stage_id"]}, {"_id": 0})
    field_values = await db.field_values.find({"card_id": card_id}, {"_id": 0}).to_list(100)
    products = await db.card_products.find({"card_id": card_id}, {"_id": 0}).to_list(100)
    history = await db.card_history.find({"card_id": card_id}, {"_id": 0}).sort("created_at", -1).to_list(20)

    # Build context for fields with their labels
    fields_info = []
    if field_values:
        field_ids = [fv["field_id"] for fv in field_values]
        fields_docs = await db.fields.find({"id": {"$in": field_ids}}, {"_id": 0}).to_list(100)
        field_label_map = {f["id"]: f["label"] for f in fields_docs}
        for fv in field_values:
            label = field_label_map.get(fv["field_id"], fv["field_id"])
            if fv.get("value_json"):
                fields_info.append(f"- {label}: {fv['value_json']}")

    context = f"""Lead: {card['nome_cliente']}
Telefone: {card.get('telefone', 'N/A')}
Email: {card.get('email', 'N/A')}
Status/Temperatura: {card.get('status', 'frio')}
Fase Atual: {stage['name'] if stage else 'N/A'}
Data de Criacao: {card.get('created_at', 'N/A')}
Campos Preenchidos:
{chr(10).join(fields_info) or 'Nenhum campo preenchido'}
Produtos:
{chr(10).join(f'- {p["nome_produto"]} (x{p["quantidade"]}) R${p.get("valor_total", 0):.2f}' for p in products) or 'Nenhum produto'}
Historico Recente:
{chr(10).join(f'- {h["action"]}: {h["details"]}' for h in history[:5]) or 'Sem historico'}"""

    try:
        chat = LlmChat(
            api_key=os.environ.get("EMERGENT_LLM_KEY"),
            session_id=f"lead-summary-{card_id}-{new_id()[:8]}",
            system_message="Voce e um assistente de CRM especializado em cosmeticos e perfumaria. Gere resumos concisos e profissionais dos leads em portugues do Brasil. Maximo 3 paragrafos curtos."
        ).with_model("anthropic", "claude-sonnet-4-5-20250929")

        response = await chat.send_message(UserMessage(text=f"Gere um resumo executivo deste lead para o vendedor:\n{context}"))
        return {"summary": response}
    except Exception as e:
        logger.error(f"AI lead summary error: {e}")
        raise HTTPException(status_code=500, detail=f"Erro ao gerar resumo: {str(e)}")

@router.post("/ai/whatsapp-suggestions/{card_id}")
async def ai_whatsapp_suggestions(card_id: str, request: Request):
    ensure_ai_available()
    user = await get_commercial_user(request)
    card = await db.cards.find_one({"id": card_id, "tenant_id": user["tenant_id"]}, {"_id": 0})
    if not card:
        raise HTTPException(status_code=404, detail="Card not found")

    messages = await db.messages.find({"card_id": card_id}, {"_id": 0}).sort("created_at", -1).to_list(10)
    stage = await db.stages.find_one({"id": card["stage_id"]}, {"_id": 0})

    conversation = "\n".join(
        f"{'Vendedor' if m['sender']=='agent' else 'Cliente'}: {m['content']}"
        for m in reversed(messages)
    ) or "Sem mensagens anteriores"

    try:
        chat = LlmChat(
            api_key=os.environ.get("EMERGENT_LLM_KEY"),
            session_id=f"wa-suggestions-{card_id}-{new_id()[:8]}",
            system_message="Voce e um assistente de vendas de cosmeticos e perfumaria. Sugira 3 respostas curtas e profissionais para enviar via WhatsApp ao cliente. Cada sugestao deve ter no maximo 2 linhas. Responda SOMENTE em JSON valido: {\"suggestions\": [\"msg1\", \"msg2\", \"msg3\"]}"
        ).with_model("anthropic", "claude-sonnet-4-5-20250929")

        response = await chat.send_message(UserMessage(
            text=f"Lead: {card['nome_cliente']}\nFase: {stage['name'] if stage else 'N/A'}\nTemperatura: {card['status']}\n\nConversa recente:\n{conversation}\n\nSugira 3 respostas profissionais para o vendedor enviar ao cliente."
        ))

        try:
            # Try to parse JSON from response
            clean = response.strip()
            if clean.startswith("```"):
                clean = clean.split("\n", 1)[1].rsplit("```", 1)[0].strip()
            parsed = json.loads(clean)
            return parsed
        except Exception:
            return {"suggestions": [response.strip()]}
    except Exception as e:
        logger.error(f"AI WhatsApp suggestions error: {e}")
        raise HTTPException(status_code=500, detail=f"Erro ao gerar sugestoes: {str(e)}")

# ============ PDF PROPOSAL GENERATION ============

@router.get("/cards/{card_id}/proposal-pdf")
async def generate_proposal_pdf(card_id: str, request: Request):
    user = await get_commercial_user(request)
    card = await db.cards.find_one({"id": card_id, "tenant_id": user["tenant_id"]}, {"_id": 0})
    if not card:
        raise HTTPException(status_code=404, detail="Card not found")

    products = await db.card_products.find({"card_id": card_id}, {"_id": 0}).to_list(100)
    tenant = await db.tenants.find_one({"id": user["tenant_id"]}, {"_id": 0})
    stage = await db.stages.find_one({"id": card["stage_id"]}, {"_id": 0})

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=30*mm, bottomMargin=20*mm, leftMargin=20*mm, rightMargin=20*mm)

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('CustomTitle', parent=styles['Title'], fontSize=24, spaceAfter=6, textColor=rl_colors.HexColor('#0A0A0B'))
    heading_style = ParagraphStyle('CustomHeading', parent=styles['Heading2'], fontSize=14, spaceAfter=8, spaceBefore=16, textColor=rl_colors.HexColor('#0A0A0B'))
    normal_style = ParagraphStyle('CustomNormal', parent=styles['Normal'], fontSize=10, spaceAfter=4, leading=14)
    small_style = ParagraphStyle('Small', parent=styles['Normal'], fontSize=8, textColor=rl_colors.HexColor('#737373'))

    elements = []

    elements.append(Paragraph("Proposta Comercial", title_style))
    org_name = tenant['name'] if tenant else 'CRM Kuryos'
    elements.append(Paragraph(org_name, ParagraphStyle('Org', parent=styles['Normal'], fontSize=12, textColor=rl_colors.HexColor('#737373'))))
    elements.append(Spacer(1, 6*mm))

    elements.append(Paragraph("Dados do Cliente", heading_style))
    elements.append(Paragraph(f"<b>Nome:</b> {card['nome_cliente']}", normal_style))
    elements.append(Paragraph(f"<b>Telefone:</b> {card.get('telefone', 'N/A')}", normal_style))
    elements.append(Paragraph(f"<b>Email:</b> {card.get('email', 'N/A')}", normal_style))
    elements.append(Paragraph(f"<b>Fase:</b> {stage['name'] if stage else 'N/A'}", normal_style))
    elements.append(Paragraph(f"<b>Data:</b> {datetime.now(timezone.utc).strftime('%d/%m/%Y')}", normal_style))
    elements.append(Spacer(1, 6*mm))

    if products:
        elements.append(Paragraph("Produtos / Servicos", heading_style))
        table_data = [["Produto", "SKU", "Qtd", "V. Unit.", "V. Total"]]
        total = 0
        for p in products:
            vt = p.get("valor_total", 0)
            total += vt
            table_data.append([
                p["nome_produto"], p.get("sku", "—"),
                str(p["quantidade"]), f"R$ {p['valor_unitario']:.2f}", f"R$ {vt:.2f}"
            ])
        table_data.append(["", "", "", "TOTAL:", f"R$ {total:.2f}"])

        table = Table(table_data, colWidths=[60*mm, 25*mm, 18*mm, 30*mm, 30*mm])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), rl_colors.HexColor('#0A0A0B')),
            ('TEXTCOLOR', (0, 0), (-1, 0), rl_colors.white),
            ('ALIGN', (2, 0), (-1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('TOPPADDING', (0, 0), (-1, 0), 8),
            ('GRID', (0, 0), (-1, -2), 0.5, rl_colors.HexColor('#E5E5E5')),
            ('FONTNAME', (3, -1), (-1, -1), 'Helvetica-Bold'),
            ('TOPPADDING', (0, -1), (-1, -1), 12),
            ('LINEABOVE', (3, -1), (-1, -1), 1, rl_colors.HexColor('#0A0A0B')),
        ]))
        elements.append(table)
    else:
        elements.append(Paragraph("Nenhum produto adicionado a esta proposta.", normal_style))

    elements.append(Spacer(1, 25*mm))
    elements.append(Paragraph("___________________________________", normal_style))
    elements.append(Paragraph(f"{user['name']} — {org_name}", normal_style))
    elements.append(Spacer(1, 8*mm))
    elements.append(Paragraph(f"Documento gerado em {datetime.now(timezone.utc).strftime('%d/%m/%Y %H:%M')} UTC", small_style))

    doc.build(elements)
    buffer.seek(0)

    filename = f"proposta_{card['nome_cliente'].replace(' ', '_')}.pdf"
    return StreamingResponse(buffer, media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'})

# ============ EXCEL REPORT EXPORT ============

@router.get("/reports/excel")
async def export_excel_report(request: Request):
    user = await get_current_user(request)
    tid = user["tenant_id"]

    wb = Workbook()
    header_fill = PatternFill(start_color="0A0A0B", end_color="0A0A0B", fill_type="solid")
    header_font = XlFont(color="FFFFFF", bold=True, size=11)

    # Sheet 1: Pipeline Funnel
    ws1 = wb.active
    ws1.title = "Funil de Vendas"
    ws1.append(["Fase", "Quantidade de Leads", "Ordem"])
    for cell in ws1[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    pipeline = await db.pipelines.find_one({"tenant_id": tid}, {"_id": 0})
    if pipeline:
        stages = await db.stages.find({"pipeline_id": pipeline["id"]}, {"_id": 0}).sort("order", 1).to_list(100)
        for s in stages:
            count = await db.cards.count_documents({"tenant_id": tid, "stage_id": s["id"]})
            ws1.append([s["name"], count, s["order"]])
    ws1.column_dimensions['A'].width = 30
    ws1.column_dimensions['B'].width = 22

    # Sheet 2: Leads
    ws2 = wb.create_sheet("Leads")
    ws2.append(["Cliente", "Telefone", "Email", "Status", "Fase", "Vendedor", "Criado em"])
    for cell in ws2[1]:
        cell.fill = header_fill
        cell.font = header_font

    cards = await db.cards.find({"tenant_id": tid}, {"_id": 0}).to_list(5000)
    all_stages = await db.stages.find({}, {"_id": 0}).to_list(100)
    stages_map = {s["id"]: s["name"] for s in all_stages}
    all_users = await db.users.find({"tenant_id": tid}, {"_id": 0, "password_hash": 0}).to_list(100)
    users_map = {u["id"]: u["name"] for u in all_users}

    for c in cards:
        ws2.append([
            c["nome_cliente"], c.get("telefone", ""), c.get("email", ""),
            c.get("status", "frio"), stages_map.get(c["stage_id"], ""),
            users_map.get(c.get("vendedor_id", ""), ""), c.get("created_at", "")
        ])
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
        ws2.column_dimensions[col].width = 22

    # Sheet 3: Tasks
    ws3 = wb.create_sheet("Tarefas")
    ws3.append(["Titulo", "Status", "Vencimento", "Criado em"])
    for cell in ws3[1]:
        cell.fill = header_fill
        cell.font = header_font

    tasks = await db.tasks.find({"tenant_id": tid}, {"_id": 0}).to_list(500)
    for t in tasks:
        ws3.append([t["titulo"], t.get("status", ""), t.get("data_vencimento", ""), t.get("created_at", "")])
    for col in ['A', 'B', 'C', 'D']:
        ws3.column_dimensions[col].width = 25

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    fname = f"relatorio_kuryos_{datetime.now(timezone.utc).strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'})

# ============ DASHBOARD ============

@router.get("/dashboard/metrics")
async def dashboard_metrics(request: Request):
    user = await get_current_user(request)
    tid = user["tenant_id"]

    total_cards = await db.cards.count_documents({"tenant_id": tid})
    cards_by_status = {}
    for status in ["frio", "morno", "quente"]:
        cards_by_status[status] = await db.cards.count_documents({"tenant_id": tid, "status": status})

    stages = await db.stages.find({}, {"_id": 0}).sort("order", 1).to_list(100)
    pipeline = await db.pipelines.find_one({"tenant_id": tid}, {"_id": 0})
    pipeline_id = pipeline["id"] if pipeline else None

    cards_by_stage = []
    for s in stages:
        count = await db.cards.count_documents({"tenant_id": tid, "stage_id": s["id"]})
        cards_by_stage.append({"stage": s["name"], "count": count, "order": s["order"]})

    total_tasks = await db.tasks.count_documents({"tenant_id": tid})
    pending_tasks = await db.tasks.count_documents({"tenant_id": tid, "status": "pendente"})
    completed_tasks = await db.tasks.count_documents({"tenant_id": tid, "status": "concluida"})

    recent_history = await db.card_history.find(
        {"card_id": {"$in": [c["id"] for c in await db.cards.find({"tenant_id": tid}, {"id": 1, "_id": 0}).to_list(5000)]}},
        {"_id": 0}
    ).sort("created_at", -1).to_list(10)

    products = await db.card_products.find(
        {"card_id": {"$in": [c["id"] for c in await db.cards.find({"tenant_id": tid}, {"id": 1, "_id": 0}).to_list(5000)]}},
        {"_id": 0}
    ).to_list(10000)
    total_revenue = sum(p.get("valor_total", 0) for p in products)

    return {
        "total_cards": total_cards, "cards_by_status": cards_by_status,
        "cards_by_stage": cards_by_stage, "total_tasks": total_tasks,
        "pending_tasks": pending_tasks, "completed_tasks": completed_tasks,
        "total_revenue": total_revenue, "recent_history": recent_history,
        "pipeline_id": pipeline_id
    }

# ============ ERP OVERVIEW (global dashboard) ============

@router.get("/erp-overview")
async def erp_overview(request: Request):
    """Aggregated KPIs from all ERP modules for the global dashboard."""
    user = await get_current_user(request)
    tid = user["tenant_id"]
    now = now_iso()

    import asyncio
    (
        total_clientes, clientes_ativos, total_projetos, projetos_ativos,
        projetos_pedido_aprovado, total_amostras, amostras_andamento,
        pd_total, pd_open, pd_in_progress, pd_in_tests, pd_waiting,
        pd_approved, pd_completed, total_formulas,
        kickoffs_preenchimento, kickoffs_aguardando, kickoffs_aprovados,
        ras_pendentes, ras_aprovadas, rncs_abertas, retencoes_ativas, checklists_pendentes,
        fornecedores_homologados, fornecedores_em_avaliacao, pos_abertas, pos_atrasadas,
        pedidos_abertos, pedidos_em_producao,
        dup_abertas, dup_vencidas,
        nfs_rascunho, recebimento_pendentes,
        contratos_total,
        wf_tasks_pendentes, wf_tasks_atrasadas,
    ) = await asyncio.gather(
        # CRM
        db.crm_clients.count_documents({"tenant_id": tid}),
        db.crm_clients.count_documents({"tenant_id": tid, "stage": {"$nin": ["cliente_perdido"]}}),
        db.crm_projects.count_documents({"tenant_id": tid}),
        db.crm_projects.count_documents({"tenant_id": tid, "stage": {"$nin": ["projeto_arquivado", "projeto_cancelado"]}}),
        db.crm_projects.count_documents({"tenant_id": tid, "stage": "pedido_aprovado"}),
        db.crm_samples.count_documents({"tenant_id": tid}),
        db.crm_samples.count_documents({"tenant_id": tid, "stage": {"$in": ["solicitada", "em_elaboracao", "retrabalho", "enviada"]}}),
        # P&D
        db.pd_requests.count_documents({"tenant_id": tid}),
        db.pd_requests.count_documents({"tenant_id": tid, "status": "OPEN"}),
        db.pd_requests.count_documents({"tenant_id": tid, "status": "IN_PROGRESS"}),
        db.pd_requests.count_documents({"tenant_id": tid, "status": "IN_TESTS"}),
        db.pd_requests.count_documents({"tenant_id": tid, "status": "WAITING_APPROVAL"}),
        db.pd_requests.count_documents({"tenant_id": tid, "status": "APPROVED"}),
        db.pd_requests.count_documents({"tenant_id": tid, "status": "COMPLETED"}),
        db.pd_formulas.count_documents({"tenant_id": tid}),
        # Kickoffs
        db.kickoffs.count_documents({"tenant_id": tid, "status": "em_preenchimento"}),
        db.kickoffs.count_documents({"tenant_id": tid, "status": "aguardando_aprovacao"}),
        db.kickoffs.count_documents({"tenant_id": tid, "status": "aprovado"}),
        # CQ
        db.cq_registros_analise.count_documents({"tenant_id": tid, "status": {"$in": ["rascunho", "em_analise"]}}),
        db.cq_registros_analise.count_documents({"tenant_id": tid, "status": "aprovado"}),
        db.cq_rncs.count_documents({"tenant_id": tid, "status": {"$in": ["aberta", "em_tratamento"]}}),
        db.cq_retencoes.count_documents({"tenant_id": tid, "status": "em_guarda"}),
        db.cq_checklists.count_documents({"tenant_id": tid, "status": {"$in": ["pendente", "em_andamento"]}}),
        # Compras
        db.compras_fornecedores.count_documents({"tenant_id": tid, "homologacao.status": "homologado"}),
        db.compras_fornecedores.count_documents({"tenant_id": tid, "homologacao.status": "em_avaliacao"}),
        db.compras_pos.count_documents({"tenant_id": tid, "status": {"$in": ["emitida", "confirmada"]}}),
        db.compras_pos.count_documents({"tenant_id": tid, "status": {"$in": ["emitida", "confirmada"]}, "data_entrega_prevista": {"$lt": now}}),
        # Pedidos
        db.orders.count_documents({"tenant_id": tid, "status": {"$in": ["rascunho", "aprovado", "em_producao"]}}),
        db.orders.count_documents({"tenant_id": tid, "status": "em_producao"}),
        # Faturamento
        db.faturamento_duplicatas.count_documents({"tenant_id": tid, "status": "aberta"}),
        db.faturamento_duplicatas.count_documents({"tenant_id": tid, "status": "vencida"}),
        db.faturamento_notas.count_documents({"tenant_id": tid, "status": "rascunho"}),
        # Recebimento
        db.recebimento_notas.count_documents({"tenant_id": tid, "status": {"$in": ["pendente", "em_conferencia"]}}),
        # Contratos
        db.contratos.count_documents({"tenant_id": tid}),
        # Workflow tasks
        db.workflow_tasks.count_documents({"tenant_id": tid, "status": {"$in": ["pendente", "em_andamento", "em_atraso"]}}),
        db.workflow_tasks.count_documents({"tenant_id": tid, "status": "em_atraso"}),
    )

    # Faturamento financial aggregations
    dup_agg = await db.faturamento_duplicatas.aggregate([
        {"$match": {"tenant_id": tid, "status": {"$in": ["aberta", "vencida"]}}},
        {"$group": {"_id": None,
            "total_em_aberto": {"$sum": "$valor"},
            "total_vencido": {"$sum": {"$cond": [{"$eq": ["$status", "vencida"]}, "$valor", 0]}},
        }},
    ]).to_list(1)
    fat_em_aberto = round((dup_agg[0]["total_em_aberto"] if dup_agg else 0), 2)
    fat_vencido = round((dup_agg[0]["total_vencido"] if dup_agg else 0), 2)

    return {
        "crm": {
            "total_clientes": total_clientes,
            "clientes_ativos": clientes_ativos,
            "total_projetos": total_projetos,
            "projetos_ativos": projetos_ativos,
            "projetos_pedido_aprovado": projetos_pedido_aprovado,
            "total_amostras": total_amostras,
            "amostras_andamento": amostras_andamento,
        },
        "pd": {
            "total": pd_total,
            "open": pd_open,
            "in_progress": pd_in_progress,
            "in_tests": pd_in_tests,
            "waiting_approval": pd_waiting,
            "approved": pd_approved,
            "completed": pd_completed,
            "formulas": total_formulas,
            "ativos": pd_in_progress + pd_in_tests + pd_waiting,
        },
        "kickoffs": {
            "em_preenchimento": kickoffs_preenchimento,
            "aguardando_aprovacao": kickoffs_aguardando,
            "aprovados": kickoffs_aprovados,
            "total": kickoffs_preenchimento + kickoffs_aguardando + kickoffs_aprovados,
        },
        "cq": {
            "ras_pendentes": ras_pendentes,
            "ras_aprovadas": ras_aprovadas,
            "rncs_abertas": rncs_abertas,
            "retencoes_ativas": retencoes_ativas,
            "checklists_pendentes": checklists_pendentes,
        },
        "compras": {
            "fornecedores_homologados": fornecedores_homologados,
            "fornecedores_em_avaliacao": fornecedores_em_avaliacao,
            "pos_abertas": pos_abertas,
            "pos_atrasadas": pos_atrasadas,
        },
        "pedidos": {
            "abertos": pedidos_abertos,
            "em_producao": pedidos_em_producao,
        },
        "faturamento": {
            "duplicatas_abertas": dup_abertas,
            "duplicatas_vencidas": dup_vencidas,
            "nfs_rascunho": nfs_rascunho,
            "total_em_aberto": fat_em_aberto,
            "total_vencido": fat_vencido,
        },
        "recebimento": {
            "pendentes": recebimento_pendentes,
        },
        "contratos": {
            "total": contratos_total,
        },
        "tasks": {
            "pendentes": wf_tasks_pendentes,
            "atrasadas": wf_tasks_atrasadas,
        },
    }


# ============ USERS (for tenant) ============

@router.get("/users")
async def list_users(request: Request):
    user = await get_current_user(request)
    users = await db.users.find({"tenant_id": user["tenant_id"]}, {"_id": 0, "password_hash": 0}).to_list(100)
    return users

# ============ SEED DEFAULT PIPELINE ============

SEED_STAGES = [
    {
        "name": "Prospeccao - Projeto", "order": 0,
        "fields": [
            {"label": "Origem da Oportunidade", "type": "select", "required": True, "options": ["Indicacao", "Site", "Evento", "LinkedIn", "Outro"]},
            {"label": "Data Inicio Prospeccao", "type": "date", "required": True, "options": []},
            {"label": "Prioridade", "type": "select", "required": True, "options": ["Alta", "Media", "Baixa"]},
            {"label": "Notas Adicionais", "type": "textarea", "required": False, "options": []},
        ]
    },
    {
        "name": "Reuniao / Contato Inicial", "order": 1,
        "fields": [
            {"label": "Qualificacao", "type": "select", "required": True, "options": ["Alto", "Medio", "Baixo"]},
            {"label": "Data da Reuniao", "type": "date", "required": True, "options": []},
            {"label": "Criterios de Qualificacao", "type": "textarea", "required": False, "options": []},
            {"label": "Documentos", "type": "file", "required": False, "options": []},
        ]
    },
    {
        "name": "Amostras", "order": 2,
        "fields": [
            {"label": "Data Solicitacao Amostra", "type": "date", "required": True, "options": []},
            {"label": "Data Envio Amostra", "type": "date", "required": False, "options": []},
            {"label": "Descricao das Amostras", "type": "textarea", "required": False, "options": []},
            {"label": "Anexos", "type": "file", "required": False, "options": []},
        ]
    },
    {
        "name": "Negociando Proposta", "order": 3,
        "fields": [
            {"label": "Proposta Enviada", "type": "boolean", "required": False, "options": []},
            {"label": "Data Envio Proposta", "type": "date", "required": False, "options": []},
            {"label": "Quantidade do Pedido", "type": "number", "required": False, "options": []},
            {"label": "Valor Unitario", "type": "number", "required": False, "options": []},
            {"label": "Valor Total", "type": "number", "required": False, "options": []},
            {"label": "Feedback do Cliente", "type": "textarea", "required": False, "options": []},
            {"label": "Anexos", "type": "file", "required": False, "options": []},
        ]
    },
    {
        "name": "Negocio Fechado", "order": 4,
        "fields": [
            {"label": "Data de Fechamento", "type": "date", "required": True, "options": []},
            {"label": "Valor Final", "type": "number", "required": True, "options": []},
        ]
    },
    {
        "name": "Negocio Perdido", "order": 5,
        "fields": [
            {"label": "Motivo da Perda", "type": "textarea", "required": True, "options": []},
        ]
    },
]

async def seed_default_pipeline(tenant_id: str):
    pipeline_id = new_id()
    await db.pipelines.insert_one({
        "id": pipeline_id, "tenant_id": tenant_id, "name": "Pipeline Padrão", "created_at": now_iso()
    })

    for stage_data in SEED_STAGES:
        stage_id = new_id()
        await db.stages.insert_one({
            "id": stage_id, "pipeline_id": pipeline_id, "name": stage_data["name"],
            "order": stage_data["order"], "created_at": now_iso()
        })
        for field_data in stage_data["fields"]:
            await db.fields.insert_one({
                "id": new_id(), "stage_id": stage_id, "label": field_data["label"],
                "type": field_data["type"], "required": field_data["required"],
                "options": field_data.get("options", []), "created_at": now_iso()
            })

    logger.info(f"Seeded default pipeline for tenant {tenant_id}")

# ============ ADMIN SEED ============

async def seed_admin():
    admin_email = os.environ.get("ADMIN_EMAIL", "admin@kuryos.com")
    admin_password = os.environ.get("ADMIN_PASSWORD", "admin123")

    existing = await db.users.find_one({"email": admin_email})
    if existing is None:
        tenant_id = new_id()
        await db.tenants.insert_one({"id": tenant_id, "name": "Kuryos Demo", "created_at": now_iso()})

        user_id = new_id()
        await db.users.insert_one({
            "id": user_id, "email": admin_email, "password_hash": hash_password(admin_password),
            "name": "Admin Kuryos", "role": "admin", "tenant_id": tenant_id, "created_at": now_iso()
        })

        await seed_default_pipeline(tenant_id)
        logger.info(f"Seeded admin user: {admin_email}")
    else:
        tenant_id = existing["tenant_id"]
        if not verify_password(admin_password, existing["password_hash"]):
            await db.users.update_one({"email": admin_email}, {"$set": {"password_hash": hash_password(admin_password)}})
            logger.info("Updated admin password")

    # Seed 8 RBAC profile users (Section 10 PRD) — one per role
    role_users = [
        ("vendedor@kuryos.com",            "Vendedor SDR",           "vendedor"),
        ("salesops@kuryos.com",            "Sales Ops",              "sales_ops"),
        ("formulador@kuryos.com",          "Formulador P&D",         "formulador"),
        ("qa@kuryos.com",                  "Controle de Qualidade",  "qa"),
        ("liderpd@kuryos.com",             "Lider P&D",              "lider_pd"),
        ("engenharia@kuryos.com",          "Engenharia de Produto",  "engenharia_produto"),
        ("sucesso@kuryos.com",             "Sucesso do Cliente",     "sucesso_cliente"),
    ]
    role_password = os.environ.get("ROLE_USERS_PASSWORD", "kuryos123")
    seeded_credentials = [(admin_email, admin_password, "admin", "Admin Kuryos")]
    for email, name, role in role_users:
        existing_user = await db.users.find_one({"email": email})
        if not existing_user:
            await db.users.insert_one({
                "id": new_id(),
                "email": email,
                "password_hash": hash_password(role_password),
                "name": name,
                "role": role,
                "tenant_id": tenant_id,
                "created_at": now_iso(),
            })
            logger.info(f"Seeded role user: {email} ({role})")
        else:
            updates = {}
            if existing_user.get("role") != role:
                updates["role"] = role
            if not verify_password(role_password, existing_user["password_hash"]):
                updates["password_hash"] = hash_password(role_password)
            if updates:
                await db.users.update_one({"email": email}, {"$set": updates})
        seeded_credentials.append((email, role_password, role, name))

    # Write credentials
    _memory_dir = Path(os.environ.get("MEMORY_DIR", str(Path(__file__).parent / "memory")))
    _memory_dir.mkdir(exist_ok=True)
    creds_md = ["# Test Credentials\n", "All users belong to the same tenant (Kuryos Demo).\n"]
    creds_md.append("## Login Endpoint\nPOST /api/auth/login  →  body: { email, password }\n")
    creds_md.append("## Users (one per RBAC profile, Section 10 PRD)\n")
    creds_md.append("| Email | Senha | Perfil | Nome |")
    creds_md.append("| --- | --- | --- | --- |")
    for email, pwd, role, name in seeded_credentials:
        creds_md.append(f"| {email} | {pwd} | {role} | {name} |")
    creds_md.append("")
    creds_md.append("## Auth Endpoints")
    creds_md.append("- POST /api/auth/login")
    creds_md.append("- POST /api/auth/register")
    creds_md.append("- GET /api/auth/me")
    creds_md.append("- POST /api/auth/logout")
    creds_md.append("- POST /api/auth/refresh")
    creds_md.append("")
    with open(_memory_dir / "test_credentials.md", "w", encoding="utf-8") as f:
        f.write("\n".join(creds_md))

# ============ STARTUP ============

@app.on_event("startup")
async def startup():
    await db.users.create_index("email", unique=True)
    await db.users.create_index("tenant_id")
    await db.cards.create_index([("tenant_id", 1), ("pipeline_id", 1)])
    await db.cards.create_index([("tenant_id", 1), ("stage_id", 1)])
    await db.stages.create_index("pipeline_id")
    await db.fields.create_index("stage_id")
    await db.field_values.create_index("card_id")
    await db.card_products.create_index("card_id")
    await db.card_history.create_index("card_id")
    await db.tasks.create_index("tenant_id")
    await db.messages.create_index("card_id")
    await db.notifications.create_index([("tenant_id", 1), ("user_id", 1)])
    await db.files.create_index("tenant_id")
    await db.email_logs.create_index("tenant_id")
    try:
        init_storage()
    except Exception as e:
        logger.warning(f"Storage init deferred: {e}")
    # P&D module indexes
    await db.pd_requests.create_index("tenant_id")
    await db.pd_requests.create_index([("tenant_id", 1), ("status", 1)])
    await db.pd_request_status_history.create_index("pd_request_id")
    await db.pd_developments.create_index("pd_request_id")
    await db.pd_developments.create_index("tenant_id")
    await db.pd_formulas.create_index("development_id")
    await db.pd_formula_items.create_index("formula_id")
    await db.pd_tests.create_index("development_id")
    await db.pd_samples.create_index("development_id")
    await db.pd_approvals.create_index("development_id")
    await db.pd_costs.create_index("development_id")
    await db.pd_documents.create_index("development_id")

    # P&D Extensions: catalog, stock, updates/pending
    await db.pd_catalog.create_index([("tenant_id", 1), ("nome", 1)])
    await db.pd_catalog.create_index([("tenant_id", 1), ("categoria", 1)])
    await db.pd_catalog_price_history.create_index("catalog_item_id")
    await db.pd_stock_items.create_index([("tenant_id", 1), ("categoria", 1)])
    await db.pd_stock_items.create_index([("tenant_id", 1), ("nome", 1)])
    await db.pd_stock_movements.create_index([("stock_item_id", 1), ("created_at", -1)])
    await db.pd_stock_movements.create_index("tenant_id")
    await db.pd_updates.create_index([("tenant_id", 1), ("pd_request_id", 1), ("created_at", -1)])
    await db.pd_pending_items.create_index([("tenant_id", 1), ("pd_request_id", 1), ("status", 1)])
    
    # Initialize P&D module
    init_pd(db, get_current_user, new_id, now_iso, put_object, broadcast_event_fn=ws_manager.broadcast)
    
    # Initialize CRM module
    init_crm(db, get_current_user, new_id, now_iso, ws_manager.broadcast)

    # Initialize Estoque module
    init_estoque(db, get_current_user, new_id, now_iso)

    # Initialize Recebimento module
    init_recebimento(db, get_current_user, new_id, now_iso)

    # Initialize Retrabalho module
    init_retrabalho(db, get_current_user, new_id, now_iso)

    # Initialize Expedição + Faturamento modules
    init_expedicao(db, get_current_user, new_id, now_iso)
    init_faturamento(db, get_current_user, new_id, now_iso)
    init_pcp(db, get_current_user, new_id, now_iso)

    # Initialize Orders module
    init_orders(db, get_current_user, new_id, now_iso)
    init_kickoff(db, get_current_user, new_id, now_iso)
    init_compras(db, get_current_user, new_id, now_iso)
    await create_compras_indexes()
    init_contratos(db, get_current_user, new_id, now_iso)

    # Initialize CQ (Controle de Qualidade) module
    init_cq(db, get_current_user, new_id, now_iso, broadcast_event_fn=ws_manager.broadcast)
    await create_cq_indexes()

    await db.orders.create_index([("tenant_id", 1), ("status", 1)])
    await db.orders.create_index([("tenant_id", 1), ("pd_request_id", 1)])
    await db.orders.create_index([("tenant_id", 1), ("created_at", -1)])
    # Kickoff
    await db.kickoffs.create_index([("tenant_id", 1), ("projeto_id", 1), ("status", 1)])
    await db.kickoffs.create_index([("tenant_id", 1), ("kickoff_group_id", 1), ("versao_numero", -1)])
    await db.kickoffs.create_index([("tenant_id", 1), ("numero_kickoff", 1)])
    # Compras (Ordens de Compra) - linked to Kickoff/BOM
    await db.ordens_compra.create_index([("tenant_id", 1), ("status", 1)])
    await db.ordens_compra.create_index([("tenant_id", 1), ("kickoff_id", 1)])
    await db.ordens_compra.create_index([("tenant_id", 1), ("fornecedor_id", 1)])
    await db.ordens_compra.create_index([("tenant_id", 1), ("numero_oc", 1)], unique=True, sparse=True)
    # Contratos CGI
    await db.contratos.create_index([("tenant_id", 1), ("kickoff_id", 1)])
    await db.contratos.create_index([("tenant_id", 1), ("client_id", 1)])
    await db.contratos.create_index([("tenant_id", 1), ("numero_contrato", 1)], unique=True, sparse=True)

    # Initialize Workflow Engine + Routes (ERP v3.0)
    init_workflow(db, new_id, now_iso)
    init_workflow_routes(db, get_current_user, new_id, now_iso)

    # Initialize Cadastros (R22 Categorias, R08 Fragrâncias)
    init_categorias(db, get_current_user, new_id, now_iso)
    await create_categorias_indexes()
    init_fragrancias(db, get_current_user, new_id, now_iso)
    await create_fragrancias_indexes()
    init_materiais(db, get_current_user, new_id, now_iso)
    await create_materiais_indexes()
    init_produtos(db, get_current_user, new_id, now_iso)
    await create_produtos_indexes()
    init_propostas(db, get_current_user, new_id, now_iso)
    init_requirements(db, get_current_user)
    await create_requirements_indexes()

    # Workflow indexes
    await db.workflow_tasks.create_index([("tenant_id", 1), ("entity_type", 1), ("entity_id", 1)])
    await db.workflow_tasks.create_index([("tenant_id", 1), ("status", 1)])
    await db.workflow_tasks.create_index([("tenant_id", 1), ("responsible_id", 1), ("status", 1)])
    await db.workflow_tasks.create_index([("tenant_id", 1), ("blocking", 1), ("status", 1)])
    await db.audit_logs.create_index([("tenant_id", 1), ("timestamp", -1)])
    await db.audit_logs.create_index([("tenant_id", 1), ("entity_type", 1), ("entity_id", 1)])
    await db.audit_logs.create_index([("tenant_id", 1), ("user_id", 1)])
    await db.audit_logs.create_index([("tenant_id", 1), ("action", 1)])

    # Estoque indexes
    await db.estoque_items.create_index([("tenant_id", 1), ("setor", 1)])
    await db.estoque_items.create_index([("tenant_id", 1), ("mp_id", 1)])
    await db.estoque_items.create_index([("tenant_id", 1), ("produto_id", 1)])
    await db.estoque_items.create_index([("tenant_id", 1), ("nome", 1)])
    await db.estoque_movimentos.create_index([("tenant_id", 1), ("item_id", 1), ("created_at", -1)])
    await db.estoque_movimentos.create_index([("tenant_id", 1), ("created_at", -1)])
    await db.estoque_movimentos.create_index([("tenant_id", 1), ("setor", 1), ("tipo", 1)])

    # Homologação indexes
    await db.homologacao_mps.create_index([("tenant_id", 1), ("status", 1)])
    await db.homologacao_mps.create_index([("tenant_id", 1), ("tipo_mp", 1)])
    await db.homologacao_mps.create_index([("tenant_id", 1), ("fornecedor_id", 1)])
    await db.homologacao_mps.create_index([("tenant_id", 1), ("nome", 1)])
    await db.homologacao_fornecedores.create_index([("tenant_id", 1), ("status", 1)])
    await db.homologacao_fornecedores.create_index([("tenant_id", 1), ("razao_social", 1)])
    await db.homologacao_fornecedores.create_index([("tenant_id", 1), ("cnpj_normalized", 1)])
    
    # CRM indexes
    await db.crm_clients.create_index([("tenant_id", 1), ("stage", 1)])
    await db.crm_clients.create_index([("tenant_id", 1)])
    await db.crm_clients.create_index([("tenant_id", 1), ("cnpj_normalized", 1)])
    await db.crm_projects.create_index([("tenant_id", 1), ("cliente_id", 1)])
    await db.crm_projects.create_index([("tenant_id", 1), ("stage", 1)])
    await db.crm_samples.create_index([("tenant_id", 1), ("projeto_id", 1)])
    await db.crm_samples.create_index([("tenant_id", 1), ("cliente_id", 1)])
    await db.crm_samples.create_index([("tenant_id", 1), ("stage", 1)])
    await db.skus.create_index([("tenant_id", 1), ("status", 1)])
    await db.skus.create_index([("tenant_id", 1), ("cliente_id", 1)])
    await db.skus.create_index([("tenant_id", 1), ("codigo_interno", 1)], unique=True)
    await db.crm_alerts.create_index([("tenant_id", 1), ("status", 1)])
    await db.crm_alerts.create_index([("tenant_id", 1), ("tipo", 1)])
    await db.crm_column_configs.create_index([("tenant_id", 1), ("crm_type", 1)])
    await db.crm_field_configs.create_index([("tenant_id", 1), ("column_id", 1)])
    await db.pd_stability_studies.create_index([("tenant_id", 1), ("pd_card_id", 1)], unique=True)
    await db.pd_stability_studies.create_index([("tenant_id", 1), ("status", 1)])
    await db.pd_stability_readings.create_index([("tenant_id", 1), ("study_id", 1), ("condition_code", 1), ("day_offset", 1)], unique=True)
    
    # Faturamento indexes
    await db.faturamento_notas.create_index([("tenant_id", 1), ("status", 1)])
    await db.faturamento_notas.create_index([("tenant_id", 1), ("status_pagamento", 1)])
    await db.faturamento_notas.create_index([("tenant_id", 1), ("created_at", -1)])
    await db.faturamento_duplicatas.create_index([("tenant_id", 1), ("status", 1)])
    await db.faturamento_duplicatas.create_index([("tenant_id", 1), ("nf_id", 1)])
    await db.faturamento_duplicatas.create_index([("tenant_id", 1), ("data_vencimento", 1)])

    # Start alert scheduler
    asyncio.create_task(run_alert_scheduler())
    asyncio.create_task(run_workflow_notification_scheduler())
    asyncio.create_task(run_stability_scheduler())
    
    await seed_admin()
    logger.info("CRM Kuryos API started")

@app.on_event("shutdown")
async def shutdown():
    client.close()

# ============ INCLUDE ROUTER + CORS + WEBSOCKET ============

app.include_router(router)
app.include_router(pd_router)
app.include_router(crm_router)
app.include_router(estoque_router)
app.include_router(recebimento_router)
app.include_router(retrabalho_router)
app.include_router(expedicao_router)
app.include_router(faturamento_router)
app.include_router(pcp_router)
app.include_router(orders_router)
app.include_router(ops_router)
app.include_router(workflow_router)
app.include_router(kickoff_router)
app.include_router(compras_router)
app.include_router(contratos_router)
app.include_router(cq_router)
app.include_router(categorias_router)
app.include_router(fragrancias_router)
app.include_router(materiais_router)
app.include_router(produtos_router)
app.include_router(propostas_router)
app.include_router(requirements_router)

@app.websocket("/api/ws")
async def websocket_endpoint(websocket: WebSocket):
    # Read token from cookies (sent automatically by browser for same-origin WS)
    token = websocket.cookies.get("access_token", "")
    if not token:
        token = websocket.query_params.get("token", "")
    if not token:
        await websocket.close(code=4001)
        return
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        tenant_id = payload.get("tenant_id", "")
        if not tenant_id:
            await websocket.close(code=4001)
            return
    except Exception:
        await websocket.close(code=4001)
        return

    await ws_manager.connect(websocket, tenant_id)
    try:
        while True:
            await websocket.receive_text()
    except WebSocketDisconnect:
        ws_manager.disconnect(websocket, tenant_id)
    except Exception:
        ws_manager.disconnect(websocket, tenant_id)

def build_allowed_origins(frontend_url: str) -> list[str]:
    origins = {
        frontend_url.rstrip("/"),
        "http://localhost:3000",
        "http://127.0.0.1:3000",
    }

    parsed = urlparse(frontend_url)
    if parsed.scheme and parsed.hostname:
        alt_host = None
        if parsed.hostname == "localhost":
            alt_host = "127.0.0.1"
        elif parsed.hostname == "127.0.0.1":
            alt_host = "localhost"

        if alt_host:
            port = f":{parsed.port}" if parsed.port else ""
            origins.add(f"{parsed.scheme}://{alt_host}{port}")

    return sorted(origins)


frontend_url = os.environ.get("FRONTEND_URL", "http://localhost:3000")
cors_origins_env = os.environ.get("CORS_ORIGINS", "").strip()

if cors_origins_env == "*":
    # Allow all origins (use regex because allow_credentials=True is incompatible with allow_origins=["*"])
    app.add_middleware(
        CORSMiddleware,
        allow_origin_regex=".*",
        allow_credentials=True,
        allow_methods=["*"],
        allow_headers=["*"],
    )
elif cors_origins_env:
    # Explicit comma-separated allowlist from env
    explicit_origins = [o.strip().rstrip("/") for o in cors_origins_env.split(",") if o.strip()]
    combined = sorted(set(explicit_origins) | set(build_allowed_origins(frontend_url)))
    app.add_middleware(
        CORSMiddleware,
        allow_origins=combined,
        allow_credentials=True,
        allow_methods=["*"],
        allow_headers=["*"],
    )
else:
    # Fallback to FRONTEND_URL-derived allowlist
    app.add_middleware(
        CORSMiddleware,
        allow_origins=build_allowed_origins(frontend_url),
        allow_credentials=True,
        allow_methods=["*"],
        allow_headers=["*"],
    )

# Mount static files for uploads
from pathlib import Path as PathLib
_default_upload_dir = PathLib(__file__).parent / "uploads"
upload_dir = PathLib(os.environ.get("UPLOAD_DIR", str(_default_upload_dir)))
upload_dir.mkdir(parents=True, exist_ok=True)
app.mount("/uploads", StaticFiles(directory=str(upload_dir)), name="uploads")
